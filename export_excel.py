import sqlite3, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

conn = sqlite3.connect("D:/edu-verify/edu_verify.db")
c = conn.cursor()

wb = openpyxl.Workbook()

hdr_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
hdr_font = Font(bold=True, size=11, color="FFFFFF")

# ── Sheet 1: 国内院校-专业 ──
ws1 = wb.active
ws1.title = "院校-专业数据"
for col, h in enumerate(["学校名称", "专业名称", "数据来源", "年份"], 1):
    cell = ws1.cell(row=1, column=col, value=h)
    cell.font = hdr_font; cell.fill = hdr_fill

c.execute("SELECT school_name, major_name, source, year FROM school_majors ORDER BY school_name, major_name")
for i, row in enumerate(c.fetchall(), 2):
    for j, val in enumerate(row, 1):
        ws1.cell(row=i, column=j, value=val)

ws1.column_dimensions["A"].width = 25
ws1.column_dimensions["B"].width = 30
ws1.column_dimensions["C"].width = 20
ws1.column_dimensions["D"].width = 8

# ── Sheet 2: 国内院校统计 ──
ws2 = wb.create_sheet("院校统计")
for col, h in enumerate(["学校名称", "专业数量"], 1):
    cell = ws2.cell(row=1, column=col, value=h)
    cell.font = hdr_font; cell.fill = hdr_fill
c.execute("SELECT school_name, COUNT(*) as cnt FROM school_majors GROUP BY school_name ORDER BY cnt DESC, school_name")
for i, row in enumerate(c.fetchall(), 2):
    ws2.cell(row=i, column=1, value=row[0])
    ws2.cell(row=i, column=2, value=row[1])
ws2.column_dimensions["A"].width = 25

# ── Sheet 3: 海外认证高校 ──
ws3 = wb.create_sheet("海外认证高校")
for col, h in enumerate(["学校名称(中文)", "国家/地区", "学校名称(原文)", "数据来源", "年份", "专业校验说明"], 1):
    cell = ws3.cell(row=1, column=col, value=h)
    cell.font = hdr_font; cell.fill = hdr_fill

c.execute("SELECT school_name, country, local_name, source, year FROM overseas_schools ORDER BY country, school_name")
for i, row in enumerate(c.fetchall(), 2):
    ws3.cell(row=i, column=1, value=row[0])
    ws3.cell(row=i, column=2, value=row[1])
    ws3.cell(row=i, column=3, value=row[2])
    ws3.cell(row=i, column=4, value=row[3])
    ws3.cell(row=i, column=5, value=row[4])
    ws3.cell(row=i, column=6, value="学校通过认证，专业以留服中心认证为准")

ws3.column_dimensions["A"].width = 30
ws3.column_dimensions["B"].width = 12
ws3.column_dimensions["C"].width = 45
ws3.column_dimensions["D"].width = 25
ws3.column_dimensions["E"].width = 8
ws3.column_dimensions["F"].width = 40

# ── Sheet 4: 海外高校统计 ──
ws4 = wb.create_sheet("海外高校统计")
for col, h in enumerate(["国家/地区", "数量"], 1):
    cell = ws4.cell(row=1, column=col, value=h)
    cell.font = hdr_font; cell.fill = hdr_fill
c.execute("SELECT country, COUNT(*) as cnt FROM overseas_schools GROUP BY country ORDER BY cnt DESC")
for i, row in enumerate(c.fetchall(), 2):
    ws4.cell(row=i, column=1, value=row[0])
    ws4.cell(row=i, column=2, value=row[1])
ws4.column_dimensions["A"].width = 15
ws4.column_dimensions["B"].width = 10

# ── 汇总 ──
ws5 = wb.create_sheet("汇总")
ws5.cell(row=1, column=1, value="数据概览").font = Font(bold=True, size=14)
c.execute("SELECT COUNT(*) FROM school_majors")
ws5.cell(row=3, column=1, value="国内院校-专业记录")
ws5.cell(row=3, column=2, value=c.fetchone()[0])
c.execute("SELECT COUNT(DISTINCT school_name) FROM school_majors")
ws5.cell(row=4, column=1, value="覆盖国内院校")
ws5.cell(row=4, column=2, value=c.fetchone()[0])
c.execute("SELECT COUNT(*) FROM overseas_schools")
ws5.cell(row=5, column=1, value="海外认证高校")
ws5.cell(row=5, column=2, value=c.fetchone()[0])
c.execute("SELECT COUNT(DISTINCT country) FROM overseas_schools")
ws5.cell(row=6, column=1, value="覆盖国家/地区")
ws5.cell(row=6, column=2, value=c.fetchone()[0])
ws5.column_dimensions["A"].width = 25

wb.save("D:/edu-verify/school_majors.xlsx")
conn.close()
print("Saved: D:/edu-verify/school_majors.xlsx")
